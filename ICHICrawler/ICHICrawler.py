from selenium import webdriver
# 공식홈페이지 Documentation 에도 ChromeOptions()가 적혀있지만 안된다
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from openpyxl import Workbook


# TODO: UnexpectedAlertPresentException: Alert Text: Error: Generic error; Means: ; Target: AV1 - 1; Action: 3-Managing > TE; No data found!
# ERROR: Alert Text: Ajax..


def getNth(prnt_li_index: int):  # Get CSS Selector(Nth of Current List Item) after tag
    # index 는 1부터 시작
    return ':nth-of-type(' + str(prnt_li_index) + ')'


def getAriaLevel(aria_depth: int):  # Get CSS Selector(Child Li List)
    return ' li[aria-level="' + str(aria_depth) + '"]'


def getWorkbook(sheet_name: str):
    work_book = Workbook()
    work_sheet = work_book.create_sheet(sheet_name)

    return work_book, work_sheet


def clickMenu(btn_menu: WebElement, target_menu: str):  # menu: Interventions, Target, Action, Means, ExtensionCodes
    if 'open' not in btn_menu.get_attribute('class'):  # 아직 열리지 않은 메뉴인 경우 클릭
        webdriver.ActionChains(driver).click(btn_menu).perform()

    trg_menu = btn_menu.find_element(By.CSS_SELECTOR, value='li[data-tsel="' + target_menu + '"]')

    webdriver.ActionChains(driver).click(trg_menu).perform()
    print("Menu " + target_menu + " Clicked")


def getFieldList(key):  # key: Interventions, Target, Action, Means, ExtensionCodes
    fields = {'Interventions': ['ICHI code', 'Target', 'Action', 'Means', 'ICHI descriptor', 'Definition',
                                'Index Terms', 'Includes Notes', 'Code also', 'Excludes Notes'],
              'Target': ['ICHI code', 'Target', 'Action', 'Means', 'Title', 'ICHI descriptor', 'Definition',
                         'Index Terms', 'Excludes', 'Includes Notes', 'Code also', 'Excludes Notes', 'Mapping ICF'],
              'Action': ['ICHI code', 'Target', 'Action', 'Means', 'Title', 'ICHI descriptor', 'Definition',
                         'Index Terms', 'Excludes', 'Includes Notes', 'Code also', 'Excludes Notes'],
              'Means': ['ICHI code', 'Target', 'Action', 'Means', 'Title', 'ICHI descriptor', 'Definition',
                        'Index Terms', 'Excludes', 'Includes Notes', 'Code also', 'Excludes Notes'],
              'ExtensionCodes': ['XCode', 'Title', 'Definition', 'Index Terms', 'Excludes', 'Notes',
                                 'Original ICD uri']}
    return fields.get(key)


def getFieldDict(field_list: list[str]):
    return dict(zip(field_list, range(len(field_list))))


def getFieldNum(field_dict: dict, key: str):
    # print("key:", key)
    num = field_dict.get(key)
    return num


def setWorkBook(sheet_name: str):
    work_book = Workbook()
    work_sheet = work_book.create_sheet(sheet_name)
    field_list = getFieldList(sheet_name)
    work_sheet.append(field_list)
    field_dict = getFieldDict(field_list)
    return work_book, work_sheet, field_dict


def appendRecord(tr_tuple: tuple[WebElement]):
    global wb, ws, f_dict
    # Excel; 필드값을 비교해 같은 곳에 value 출력
    # print("size:", len(f_dict))
    record = ['' for _ in range(len(f_dict))]
    for tr in tr_tuple:
        field = tr.find_element(By.CSS_SELECTOR, value="td:nth-of-type(1)").text
        field_idx = getFieldNum(f_dict, field)
        # print(field_idx)
        data = tr.find_element(By.CSS_SELECTOR, value="td:nth-of-type(2)").text
        record[field_idx] = data

    ws.append(record)


def DFS(prnt_tuple: tuple[WebElement], prnt_selector: str, prnt_depth: int):
    global wb, xlsx_file_name
    li_nth = 1
    tab: str = '   ' * (prnt_depth - 1) + '+  '

    # print(tab + "* * * * * * * Lv." + str(prnt_depth) + " DFS Start * * * * * * *")
    # print(tab + prnt_selector)  # cur item List

    for li in prnt_tuple:
        # get anchor
        trg_anchor = li.find_element(By.TAG_NAME, value='a')
        if trg_anchor == '#':  # TODO try 해서 경고문이 뜨면 catch 알려주고 넘어간다
            print("No Data Found!")
            li_nth = li_nth + 1
            continue
        else:
            webdriver.ActionChains(driver).double_click(trg_anchor).perform()
            # get nth List Item
            nth_li_elmt = driver.find_element(By.CSS_SELECTOR, value=prnt_selector + getNth(li_nth))
        if 'jstree-leaf' in nth_li_elmt.get_attribute('class'):
            # Get Right Contents(leaf format)
            print(tab + '- ' + str(li_nth) + '. ' + trg_anchor.text)  # print a

            tr_tpl = tuple(driver.find_elements(By.CSS_SELECTOR, value="#content tbody > tr"))
            appendRecord(tr_tpl)

            li_nth = li_nth + 1
            continue
        else:
            # Get Right Contents
            print(tab + str(li_nth) + '. ' + trg_anchor.text)  # print a; step 2~6

            tr_tpl = tuple(driver.find_elements(By.CSS_SELECTOR, value="#content tbody > tr"))
            appendRecord(tr_tpl)

            next_selector: str = prnt_selector + getNth(li_nth) + getAriaLevel(prnt_depth + 1)
            child_tpl = tuple(driver.find_elements(By.CSS_SELECTOR, value=next_selector))
            DFS(child_tpl, next_selector, prnt_depth + 1)

            li_nth = li_nth + 1

        wb.save(xlsx_file_name)


# Workbook
menu = 'Action'  # set menu
xlsx_file_name = 'ICHI_Beta_3_' + menu + '.xlsx'
wb, ws, f_dict = setWorkBook(sheet_name=menu)

# TODO 파일 이름 새로 지정; 실수로 덮어쓰지 않도록 해야함; 저장 메서드 만들기;
wb.save(filename=xlsx_file_name)

# Web Crawling
mainPage = 'https://mitel.dimi.uniud.it/ichi/'

opt = Options()
opt.add_argument('headless')
opt.add_argument('--blink-settings=imagesEnabled=false')
s = Service('../BrowserDriver/chromedriver.exe')
driver = webdriver.Chrome(service=s, options=opt)
driver.implicitly_wait(20)

print("Chrome Driver Ready")
print()
driver.get(mainPage)

clickMenu(driver.find_element(By.CSS_SELECTOR, value='#dropdown-tsel'), target_menu=menu)

DFS(tuple(driver.find_elements(By.CSS_SELECTOR, value='#tree li[aria-level="1"]')), '#tree li[aria-level="1"]', 1)
wb.save(filename=xlsx_file_name)
